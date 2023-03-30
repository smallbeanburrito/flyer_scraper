import chromedriver_autoinstaller
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from food_basics import foodbasics_scrape
from nofrills import nofrills_scrape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

chromedriver_autoinstaller.install()
driver = webdriver.Chrome()

nofrills_scrape(driver)
#product_info = foodbasics_scrape(driver)

"""

ads = Workbook()
flyer_items = ads.active
# save into excel
absRow = 2
for column in product_info:
    absCol = 1
    flyer_items[get_column_letter(absCol) + str(absRow)] = column[0]
    flyer_items[get_column_letter(absCol+1) + str(absRow)] = column[1]
    absRow = absRow + 1

ads.save("flyers.xlsx")
"""







